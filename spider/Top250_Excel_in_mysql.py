from openpyxl import load_workbook
import pymysql

# 将抓取下来的电影写入mysql数据库
config = {
    'host': '127.0.0.1',
    'port': 3306,
    'user': 'root',
    'password': 'root',
    'charset': 'utf8',
    # 'cursorclass': pymysql.cursors.DictCursor

}

conn = pymysql.connect(**config)

conn.autocommit(1)

cursor = conn.cursor()

name = 'douban_top_250'

cursor.execute('create database if not exists %s CHARACTER SET utf8 COLLATE utf8_general_ci' % name)

conn.select_db(name)

table_name = 'info'

cursor.execute(
    'create table if not exists %s(m_id int(10) not null auto_increment primary key,name varchar(255),star_con varchar(255),score varchar(255),info_list varchar(255))' % table_name)

wb2 = load_workbook('mv.xlsx')

ws = wb2.sheetnames

for row in wb2:
    for cell in row:
        info_name = cell[0].value
        info_star_con = cell[1].value
        info_score = cell[2].value
        info_info_list = cell[3].value
        sql = """INSERT INTO info(name, star_con, score, info_list) VALUES ('%s', '%s', '%s', '%s' )""" % (
            info_name, info_star_con, info_score, info_info_list)
        print(sql)
        cursor.execute(sql)
